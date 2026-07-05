
import json
from pathlib import Path

from dotenv import load_dotenv
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config.settings import Settings

load_dotenv()


class Planner:

    def __init__(self):
        self.client = genai.Client(api_key=Settings.GEMINI_API_KEY)

    def load_prompt(self):
        prompt_file = (
            Path(__file__).resolve().parent.parent
            / "prompts"
            / "planner_prompt.txt"
        )
        return prompt_file.read_text(encoding="utf-8")

    def generate_test_cases(self):
        prompt = self.load_prompt()

        requirement_file = (
            Path(__file__).resolve().parent.parent
            / "requirements"
            / "login_requirement.txt"
        )

        requirement = requirement_file.read_text(encoding="utf-8")

        prompt = prompt.replace("{requirement}", requirement)

        response = self.client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config={
                "response_mime_type": "application/json"
            }
        )

        try:
            print("=" * 60)
            print("Gemini Response")
            print("=" * 60)
            print(response.text)
            return json.loads(response.text)
        except Exception:
            print("\nGemini Response:\n")
            print(response.text)
            raise

    def remove_duplicates(self, testcases):
        unique = []
        seen = set()

        for tc in testcases:
            scenario = tc.get("scenario", "").strip().lower()
            if scenario not in seen:
                seen.add(scenario)
                unique.append(tc)

        return unique

    def assign_ids(self, testcases):
        for i, tc in enumerate(testcases, start=1):
            tc["id"] = f"TC{i:02d}"
        return testcases

    def format_value(self, value):
        if value is None:
            return ""

        if isinstance(value, list):
            return "\n".join(map(str, value))

        if isinstance(value, dict):
            return json.dumps(value, indent=2)

        return str(value)

    def save_excel(self, testcases):
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        headers = [
            "ID",
            "Module",
            "Scenario",
            "Steps",
            "Test Data",
            "Expected Result",
            "Priority",
            "Test Type",
            "Execution Type",
            "Automation Status"
        ]

        ws.append(headers)

        fill = PatternFill(start_color="1F4E78",
                           end_color="1F4E78",
                           fill_type="solid")
        font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

        for tc in testcases:
            ws.append([
                tc.get("id", ""),
                tc.get("module", ""),
                tc.get("scenario", ""),
                self.format_value(tc.get("steps")),
                self.format_value(tc.get("test_data")),
                self.format_value(tc.get("expected")),
                tc.get("priority", ""),
                tc.get("test_type", ""),
                tc.get("execution_type", ""),
                tc.get("automation_status", "")
            ])

        for column in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 5, 50)

        output = Path(__file__).resolve().parent / "testcases.xlsx"
        wb.save(output)

        print(f"\n✅ Excel Created: {output}")

    def run(self):
        print("=" * 60)
        print("🚀 Agentic Planner Started")
        print("=" * 60)

        print("\nGenerating AI Test Cases...\n")

        testcases = self.generate_test_cases()
        print(f"Generated : {len(testcases)}")

        testcases = self.remove_duplicates(testcases)
        print(f"After Duplicate Removal : {len(testcases)}")

        testcases = self.assign_ids(testcases)

        self.save_excel(testcases)

        print("\n✅ Planner Completed Successfully")


if __name__ == "__main__":
    Planner().run()
