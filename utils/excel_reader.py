from openpyxl import load_workbook


class ExcelReader:

    def __init__(self, file):

        self.file = file

    def get_testcases(self, execution_type=None):

        wb = load_workbook(self.file)

        ws = wb.active

        testcases = []

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):

            tc = dict(zip(headers, row))

            if execution_type:

                if tc["Execution Type"].lower() != execution_type.lower():
                    continue

            testcases.append(tc)

        return testcases